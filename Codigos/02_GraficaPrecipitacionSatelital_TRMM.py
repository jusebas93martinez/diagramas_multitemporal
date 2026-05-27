# -*- coding: utf-8 -*-
"""
SCRIPT UNIFICADO: Análisis Completo de Precipitación Satelital (TRMM y GPM IMERG)
- Extracción de datos de rasters TRMM/GPM IMERG
- Análisis estadístico completo
- Tablas Intranual e Interanual con filtro La Niña
- Gráficas de series temporales y promedios
- Reporte TXT unificado
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.dates as mdates
import numpy as np
import os
from datetime import datetime
import rasterio
import glob
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 0. CONFIGURACIÓN INICIAL (EDITAR AQUÍ)
# ==========================================

# Nombre del sitio de análisis (aparecerá en títulos y archivos)
NOMBRE_SITIO = "CIÉNAGA LA CHIQUITA"  # Nombre del sitio para títulos y reportes

# Coordenadas del punto de análisis
LATITUD = 8.728617
LONGITUD = -75.91079

# Año de inicio para la gráfica comparativa TRMM vs IDEAM
# None  → usa todos los años disponibles desde el inicio del registro
# Ejemplo: ANIO_INICIO_COMPARACION = 2000  (grafica solo desde 2000 en adelante)
ANIO_INICIO_COMPARACION = None

# Carpeta con los rasters TRMM/GPM IMERG
RUTA_RASTERS = r"C:\Users\sebas\OneDrive\Documentos\SYNC\ANT\CODIGO_MULTITEMPORAL\precipitacion_mensual_colombia_10km"

# Carpeta de salida
RUTA_SALIDA = r"C:\Users\sebas\OneDrive\Documentos\SYNC\ANT\2026-1\20260423\LA CHIQUITA\TRMM\2"

# Ruta al TXT de análisis de la estación IDEAM (análisis combinado)
RUTA_IDEAM_TXT = r"C:\Users\sebas\OneDrive\Documentos\SYNC\ANT\2026-1\20260423\LA CHIQUITA\IDEAM\MONTERIA\descargaDhime_analisis_completo.txt"

# Ruta al CSV de datos brutos de la estación IDEAM (gráfica comparativa)
RUTA_IDEAM_CSV = r"C:\Users\sebas\OneDrive\Documentos\SYNC\ANT\2026-1\20260423\LA CHIQUITA\IDEAM\MONTERIA\descargaDhime.csv"

# Tipo de estación IDEAM para el análisis combinado con TRMM
# 'pluviometrica': valores en mm de precipitación → comparación directa con TRMM (mismo sistema de unidades)
# 'linimetrica'  : valores en mm de nivel de regla limnimétrica → comparación por anomalías
#                  normalizadas (z-scores), ya que los sistemas de medición no son comparables en
#                  unidades absolutas. Soporta registros anteriores a 1998 (pre-TRMM).
TIPO_ESTACION_IDEAM = 'linimetrica'

# Colores para las tablas
COLOR_NINA = "#CFCFCF"       # Gris para La Niña
COLOR_HUMEDO = '#ADD8E6'     # Azul claro (> P75)
COLOR_SECO = '#F08080'       # Rojo claro (< P25)
COLOR_TOP_VERDE = '#90EE90'  # Verde (mayores)
COLOR_TOP_ROJO = '#FFB6C1'   # Rosa (menores)

# ==========================================
# 1. FUNCIONES AUXILIARES
# ==========================================

def es_periodo_nina(year, month):
    """Verifica si una fecha cae dentro de los periodos de La Niña."""
    fecha_actual = datetime(year, month, 1)
    periodos_nina = [
        (datetime(1988, 6, 1), datetime(1989, 6, 1)),
        (datetime(1998, 6, 1), datetime(2001, 6, 1)),
        (datetime(2007, 8, 1), datetime(2008, 6, 1)),
        (datetime(2010, 6, 1), datetime(2011, 4, 1)),
        (datetime(2016, 10, 1), datetime(2017, 2, 1)),
        (datetime(2020, 9, 1), datetime(2023, 2, 1))
    ]
    for inicio, fin in periodos_nina:
        if inicio <= fecha_actual <= fin:
            return True
    return False

def es_anio_nina(year):
    """Determina si un año es predominantemente La Niña (>= 6 meses afectados)."""
    count = sum(1 for m in range(1, 13) if es_periodo_nina(year, m))
    return count >= 6

def agregar_leyenda(ax):
    """Agrega leyenda estándar a las tablas."""
    elementos = [
        mpatches.Patch(facecolor=COLOR_HUMEDO, edgecolor='gray', label='> P75 (Húmedo)'),
        mpatches.Patch(facecolor=COLOR_SECO, edgecolor='gray', label='< P25 (Seco)'),
        mpatches.Patch(facecolor=COLOR_TOP_VERDE, edgecolor='gray', label='Mayores'),
        mpatches.Patch(facecolor=COLOR_TOP_ROJO, edgecolor='gray', label='Menores')
    ]
    ax.legend(handles=elementos, loc='lower center', bbox_to_anchor=(0.5, 1.02),
              ncol=4, frameon=False, fontsize=9)

# ==========================================
# 2. CREAR CARPETA DE SALIDA
# ==========================================

if not os.path.exists(RUTA_SALIDA):
    os.makedirs(RUTA_SALIDA)
    print(f"Carpeta de salida creada: {RUTA_SALIDA}")

# Nombre base para archivos (sin espacios ni caracteres especiales)
nombre_base = f"TRMM_{NOMBRE_SITIO.replace(' ', '_')}"

# Rutas de archivos de salida
ruta_csv = os.path.join(RUTA_SALIDA, f"{nombre_base}.csv")
ruta_png_serie = os.path.join(RUTA_SALIDA, f"{nombre_base}_serie_temporal.png")
ruta_png_intranual = os.path.join(RUTA_SALIDA, f"{nombre_base}_tabla_intranual.png")
ruta_png_interanual = os.path.join(RUTA_SALIDA, f"{nombre_base}_tabla_interanual.png")
ruta_png_prom_anual = os.path.join(RUTA_SALIDA, f"{nombre_base}_promedio_anual.png")
ruta_png_prom_mensual = os.path.join(RUTA_SALIDA, f"{nombre_base}_promedio_mensual.png")
ruta_png_meses_humedos = os.path.join(RUTA_SALIDA, f"{nombre_base}_meses_humedos_seleccion.png")
ruta_png_heatmap = os.path.join(RUTA_SALIDA, f"{nombre_base}_heatmap.png")
ruta_analisis = os.path.join(RUTA_SALIDA, f"{nombre_base}_ANALISIS_COMPLETO.txt")
ruta_excel_regimen = os.path.join(RUTA_SALIDA, f"{nombre_base}_regimen_pluviometrico.xlsx")

print("="*80)
print(f"INICIANDO ANÁLISIS PARA: {NOMBRE_SITIO} (Lat {LATITUD}, Lon {LONGITUD})")
print("="*80)

# ==========================================
# 3. EXTRACCIÓN DE DATOS DE RASTERS
# ==========================================

archivos_tif = glob.glob(os.path.join(RUTA_RASTERS, "*.tif"))

if not archivos_tif:
    print(f"No se encontraron archivos .tif en: {RUTA_RASTERS}")
    exit()

datos_extraidos = []
print(f"Se encontraron {len(archivos_tif)} archivos. Extrayendo datos...")

for archivo in archivos_tif:
    nombre_base_archivo = os.path.basename(archivo)
    try:
        match = re.search(r'(\d{4})_(\d{2})', nombre_base_archivo)
        if match:
            año = int(match.group(1))
            mes = int(match.group(2))
        else:
            continue

        with rasterio.open(archivo) as src:
            row, col = src.index(LONGITUD, LATITUD)
            valor = src.read(1, window=rasterio.windows.Window(col, row, 1, 1))[0, 0]

            if valor == src.nodata or valor < 0:
                valor = np.nan

            datos_extraidos.append({"AÑO": año, "MES": mes, "VALOR": float(valor)})

    except Exception as e:
        print(f"Error procesando {nombre_base_archivo}: {e}")

df = pd.DataFrame(datos_extraidos)

if df.empty:
    print("No se extrajeron datos. Verifica las coordenadas o los archivos.")
    exit()

# Eliminar NaNs y valores 0
registros_antes = len(df)
df = df.dropna(subset=['VALOR'])
df = df[df['VALOR'] > 0].copy()
registros_eliminados = registros_antes - len(df)
if registros_eliminados > 0:
    print(f"Se eliminaron {registros_eliminados} registros (NoData o valor 0)")

# Excluir año actual incompleto (2025)
df = df[df['AÑO'] != 2025]

# Ordenar y guardar CSV
df = df.sort_values(by=["AÑO", "MES"]).reset_index(drop=True)
df.to_csv(ruta_csv, index=False, sep=',')
print(f"Datos extraídos y guardados en: {ruta_csv}")
print(f"Registros válidos: {len(df)}")

# ==========================================
# 4. PROCESAMIENTO ESTADÍSTICO
# ==========================================

# Crear columna de fecha
df["FECHA"] = pd.to_datetime(df["AÑO"].astype(str) + "-" + df["MES"].astype(str) + "-01")

# Media móvil 12 meses
df["MA12"] = df["VALOR"].rolling(window=12, center=True).mean()

# Variable tiempo en años
df["t"] = (df["FECHA"] - df["FECHA"].min()).dt.days / 365.25

# Tendencia lineal
y_vals = df["VALOR"].values
t_vals = df["t"].values
coef_lin = np.polyfit(t_vals, y_vals, 1)
tend_lin = np.poly1d(coef_lin)(t_vals)
pendiente = coef_lin[0]
cambio_decadal = pendiente * 10

# Estadísticas mensuales
meses_nombres = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
meses_abbr = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC']

estadisticas_mensuales = df.groupby("MES")["VALOR"].agg(['mean', 'std', 'min', 'max', 'median']).round(2)
estadisticas_mensuales['cv'] = (estadisticas_mensuales['std'] / estadisticas_mensuales['mean'] * 100).round(2)

# Estadísticas anuales
media_mensual = df["VALOR"].mean()
estadisticas_anuales = df.groupby('AÑO')['VALOR'].agg(['sum', 'mean']).round(2)

# Meses húmedos y secos
meses_humedos = estadisticas_mensuales[estadisticas_mensuales['mean'] > media_mensual].index.tolist()
meses_secos = estadisticas_mensuales[estadisticas_mensuales['mean'] <= media_mensual].index.tolist()

# Años extremos
percentil_10_anual = estadisticas_anuales['sum'].quantile(0.1)
percentil_90_anual = estadisticas_anuales['sum'].quantile(0.9)
años_secos = estadisticas_anuales[estadisticas_anuales['sum'] <= percentil_10_anual].index.tolist()
años_humedos = estadisticas_anuales[estadisticas_anuales['sum'] >= percentil_90_anual].index.tolist()

# Anomalías
df['ANOMALIA'] = df['VALOR'] - media_mensual
df['ANOMALIA_PCT'] = ((df['VALOR'] - media_mensual) / media_mensual * 100).round(2) if media_mensual > 0 else 0

# Análisis por décadas
df['DECADA'] = (df['AÑO'] // 10) * 10
estadisticas_decadas = df.groupby('DECADA')['VALOR'].agg(['mean', 'sum', 'count']).round(2)

# ==========================================
# 5. GRÁFICA SERIE TEMPORAL
# ==========================================

print("Generando gráfica de serie temporal...")

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(16, 10), height_ratios=[3, 1])

# Panel superior
ax1.bar(df["FECHA"], df["VALOR"], width=20, color='skyblue', alpha=0.5,
        label="Precipitación mensual", edgecolor='navy', linewidth=0.5)
ax1.plot(df["FECHA"], df["MA12"], color='red', linewidth=2, label="Media móvil 12 meses")
ax1.plot(df["FECHA"], tend_lin, color='green', linestyle='--', linewidth=2,
         label=f"Tendencia ({cambio_decadal:.1f} mm/década)")
ax1.axhline(y=media_mensual, color='gray', linestyle=':', linewidth=1.5,
            label=f'Media histórica: {media_mensual:.1f} mm')

# Valores extremos
percentil_95 = df["VALOR"].quantile(0.95)
extremos = df[df["VALOR"] > percentil_95]
ax1.scatter(extremos["FECHA"], extremos["VALOR"], color='red', s=50, zorder=5, label='Extremos (>P95)')

ax1.set_title(f"Análisis de Precipitación Satelital (TRMM/GPM IMERG) - {NOMBRE_SITIO}", fontsize=14, fontweight='bold')
ax1.set_ylabel("Precipitación (mm)")
ax1.grid(True, alpha=0.3, linestyle='--')
ax1.legend(loc='upper left', fontsize=10)
ax1.set_xlim(df["FECHA"].min(), df["FECHA"].max())

# Panel inferior - Anomalías porcentuales
colores = ['#d73027' if x < 0 else '#4575b4' for x in df['ANOMALIA_PCT']]
ax2.bar(df["FECHA"], df["ANOMALIA_PCT"], width=20, color=colores, alpha=0.7)
ax2.axhline(y=0, color='black', linewidth=2, label='Media histórica (0%)')

# Anotación explicativa sobre la línea cero
ax2.annotate('← Línea cero = Media histórica', xy=(df["FECHA"].min(), 0),
             xytext=(50, 30), textcoords='offset points', fontsize=8,
             arrowprops=dict(arrowstyle='->', color='gray', lw=1),
             bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor='gray', alpha=0.8))

ax2.set_xlabel("Fecha")
ax2.set_ylabel("Variación (%)")
ax2.set_title("VARIACIONES PORCENTUALES (% respecto a la media histórica)", fontsize=11, fontweight='bold')
ax2.grid(True, alpha=0.3, linestyle='--')
ax2.set_xlim(df["FECHA"].min(), df["FECHA"].max())

# Añadir leyenda pequeña para el panel de anomalías
legend_anom = [
    mpatches.Patch(facecolor='#4575b4', alpha=0.7, label='Húmedo (+%)'),
    mpatches.Patch(facecolor='#d73027', alpha=0.7, label='Seco (-%)')
]
ax2.legend(handles=legend_anom, loc='upper right', fontsize=8, framealpha=0.9)

for ax in [ax1, ax2]:
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))

plt.tight_layout()
plt.savefig(ruta_png_serie, dpi=200, bbox_inches='tight')
plt.close()
print(f"  Guardado: {ruta_png_serie}")

# ==========================================
# 6. TABLAS INTRANUAL E INTERANUAL
# ==========================================

# Crear tabla pivote
tabla = df.pivot(index="AÑO", columns="MES", values="VALOR").sort_index()
tabla.columns = meses_abbr

# --- TABLA INTRANUAL ---
print("Generando tabla intranual...")

p75_anual = tabla.quantile(0.75, axis=1)
p25_anual = tabla.quantile(0.25, axis=1)
prom_anual = tabla.mean(axis=1)

top5_p75 = p75_anual.nlargest(5).index.tolist()
bot5_p75 = p75_anual.nsmallest(5).index.tolist()
top5_prom = prom_anual.nlargest(5).index.tolist()
bot5_prom = prom_anual.nsmallest(5).index.tolist()

encabezados_intra = ["AÑO"] + meses_abbr + ["P75", "P25", "Promedio"]
celdas_texto = []
celdas_color = []

for año in tabla.index:
    fila = tabla.loc[año]
    textos = [str(año)] + [f"{v:.1f}" if pd.notnull(v) else "-" for v in fila] + \
             [f"{p75_anual[año]:.1f}", f"{p25_anual[año]:.1f}", f"{prom_anual[año]:.1f}"]

    colores = ['white']
    for i, m in enumerate(meses_abbr):
        v = fila.iloc[i]
        m_num = i + 1
        c = 'white'
        if pd.notnull(v):
            if v > p75_anual[año]:
                c = COLOR_HUMEDO
            elif v < p25_anual[año]:
                c = COLOR_SECO
        colores.append(c)

    # Columnas de estadísticas
    colores.append(COLOR_TOP_VERDE if año in top5_p75 else (COLOR_TOP_ROJO if año in bot5_p75 else 'white'))
    colores.append('white')
    colores.append(COLOR_TOP_VERDE if año in top5_prom else (COLOR_TOP_ROJO if año in bot5_prom else 'white'))

    celdas_texto.append(textos)
    celdas_color.append(colores)

altura_fila = 0.25
espacio_titulo = 1.5
num_filas = len(tabla) + 1
altura_total = (num_filas * altura_fila) + espacio_titulo

fig1, ax1 = plt.subplots(figsize=(16, altura_total))
margen_superior = 1 - (espacio_titulo / altura_total)
plt.subplots_adjust(top=margen_superior, bottom=0.01, left=0.02, right=0.98)

t1 = ax1.table(cellText=celdas_texto, cellColours=celdas_color, colLabels=encabezados_intra,
               cellLoc='center', loc='upper left', bbox=[0, 0, 1, 1])
t1.auto_set_font_size(False)
t1.set_fontsize(9)

for (i, j), cel in t1.get_celld().items():
    if i == 0:
        cel.set_facecolor('#2E75B6')
        cel.get_text().set_color('white')

ax1.axis('off')
ax1.set_title(f"ANÁLISIS INTRANUAL - {NOMBRE_SITIO}", fontsize=14, fontweight='bold', y=1.08)
agregar_leyenda(ax1)

plt.savefig(ruta_png_intranual, dpi=150, bbox_inches='tight')
plt.close(fig1)
print(f"  Guardado: {ruta_png_intranual}")

# --- TABLA INTERANUAL ---
print("Generando tabla interanual...")

p75_mens = tabla.quantile(0.75, axis=0)
p25_mens = tabla.quantile(0.25, axis=0)
prom_mens = tabla.mean(axis=0)

top3_prom_mens = prom_mens.nlargest(3).index.tolist()
bot3_prom_mens = prom_mens.nsmallest(3).index.tolist()

encabezados_inter = ["AÑO"] + meses_abbr
celdas_texto_i = []
celdas_color_i = []

for año in tabla.index:
    fila = tabla.loc[año]
    textos = [str(año)] + [f"{v:.1f}" if pd.notnull(v) else "-" for v in fila]
    colores = ['white']

    for i, m_nom in enumerate(meses_abbr):
        v = fila.iloc[i]
        m_num = i + 1
        c = 'white'
        if pd.notnull(v):
            if v > p75_mens[m_nom]:
                c = COLOR_HUMEDO
            elif v < p25_mens[m_nom]:
                c = COLOR_SECO
        colores.append(c)
    celdas_texto_i.append(textos)
    celdas_color_i.append(colores)

# Filas resumen
def colores_resumen(top, bot):
    cols = ['#E0E0E0']
    for m in meses_abbr:
        if m in top: cols.append(COLOR_TOP_VERDE)
        elif m in bot: cols.append(COLOR_TOP_ROJO)
        else: cols.append('white')
    return cols

celdas_texto_i.append(["P75 Hist."] + [f"{p75_mens[m]:.1f}" for m in meses_abbr])
celdas_color_i.append(colores_resumen(p75_mens.nlargest(3).index.tolist(), p75_mens.nsmallest(3).index.tolist()))

celdas_texto_i.append(["P25 Hist."] + [f"{p25_mens[m]:.1f}" for m in meses_abbr])
celdas_color_i.append(colores_resumen(p25_mens.nlargest(3).index.tolist(), p25_mens.nsmallest(3).index.tolist()))

celdas_texto_i.append(["Prom. Hist."] + [f"{prom_mens[m]:.1f}" for m in meses_abbr])
celdas_color_i.append(colores_resumen(top3_prom_mens, bot3_prom_mens))

num_filas_i = len(celdas_texto_i) + 1
altura_total_i = (num_filas_i * altura_fila) + espacio_titulo
fig2, ax2 = plt.subplots(figsize=(16, altura_total_i))

plt.subplots_adjust(top=1-(espacio_titulo/altura_total_i), bottom=0.01, left=0.02, right=0.98)

t2 = ax2.table(cellText=celdas_texto_i, cellColours=celdas_color_i, colLabels=encabezados_inter,
               cellLoc='center', loc='upper left', bbox=[0, 0, 1, 1])
t2.auto_set_font_size(False)
t2.set_fontsize(9)

for (i, j), cel in t2.get_celld().items():
    if i == 0:
        cel.set_facecolor('#2E75B6')
        cel.get_text().set_color('white')

ax2.axis('off')
ax2.set_title(f"ANÁLISIS INTERANUAL - {NOMBRE_SITIO}", fontsize=14, fontweight='bold', y=1.08)
agregar_leyenda(ax2)

plt.savefig(ruta_png_interanual, dpi=150, bbox_inches='tight')
plt.close(fig2)
print(f"  Guardado: {ruta_png_interanual}")

# ==========================================
# 7. GRÁFICAS DE PROMEDIOS
# ==========================================

print("Generando gráficas de promedios...")

# Promedio Anual
fig_anual, ax_anual = plt.subplots(figsize=(14, 7))

años_list = prom_anual.index.tolist()
valores_list = prom_anual.values.tolist()
media_total = prom_anual.mean()

# Sombreado suave para años La Niña (fondo, zorder bajo)
for año in años_list:
    if es_anio_nina(año):
        ax_anual.axvspan(año - 0.5, año + 0.5, alpha=0.18, color=COLOR_NINA, zorder=1)

# Línea principal + relleno bajo la curva
ax_anual.plot(años_list, valores_list, marker='o', linestyle='-', color='#2E75B6',
              linewidth=2, markersize=5, zorder=3)
ax_anual.fill_between(años_list, valores_list, color='#2E75B6', alpha=0.1)

# Media histórica
ax_anual.axhline(y=media_total, color='navy', linestyle='--', linewidth=1.5)

# Destacar top5 (verde) y bot5 (rojo) con scatter + anotación
for año in top5_prom:
    if año in prom_anual.index:
        ax_anual.scatter(año, prom_anual[año], color='green', s=100, zorder=5)
        ax_anual.annotate(f'{prom_anual[año]:.1f}', xy=(año, prom_anual[año]),
                          textcoords='offset points', xytext=(0, 10),
                          ha='center', fontsize=8, fontweight='bold', color='darkgreen')

for año in bot5_prom:
    if año in prom_anual.index:
        ax_anual.scatter(año, prom_anual[año], color='red', s=100, zorder=5)
        ax_anual.annotate(f'{prom_anual[año]:.1f}', xy=(año, prom_anual[año]),
                          textcoords='offset points', xytext=(0, -15),
                          ha='center', fontsize=8, fontweight='bold', color='darkred')

# Leyenda
leyenda_elementos = [
    plt.Line2D([0], [0], color='#2E75B6', linewidth=2, label='Promedio mensual por año'),
    plt.Line2D([0], [0], color='navy', linestyle='--', linewidth=1.5, label=f'Media histórica: {media_total:.1f} mm'),
    plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='green', markersize=9, label='5 años más húmedos'),
    plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='red', markersize=9, label='5 años más secos'),
    mpatches.Patch(facecolor=COLOR_NINA, alpha=0.4, edgecolor='gray', label='Años La Niña'),
]
ax_anual.legend(handles=leyenda_elementos, loc='upper left', fontsize=9, framealpha=0.9)

ax_anual.set_title(f"Promedio Precipitación Mensual por Año - {NOMBRE_SITIO}", fontsize=14, fontweight='bold')
ax_anual.set_xlabel("Año")
ax_anual.set_ylabel("Precipitación Promedio (mm)")
ax_anual.set_xticks(años_list)
ax_anual.set_xticklabels(años_list, rotation=45)
ax_anual.grid(True, linestyle='--', alpha=0.5)
ax_anual.set_xlim(años_list[0] - 0.5, años_list[-1] + 0.5)
plt.tight_layout()
plt.savefig(ruta_png_prom_anual, dpi=150)
plt.close()
print(f"  Guardado: {ruta_png_prom_anual}")

# Promedio Mensual
plt.figure(figsize=(10, 6))
plt.plot(meses_abbr, prom_mens.values, marker='o', linestyle='-', color='#2E75B6', linewidth=2)
plt.fill_between(meses_abbr, prom_mens.values, color='#2E75B6', alpha=0.1)

for m in top3_prom_mens:
    plt.scatter(m, prom_mens[m], color='green', s=100, zorder=5)
    plt.annotate(f"{prom_mens[m]:.1f}", (m, prom_mens[m]), textcoords="offset points",
                 xytext=(0,10), ha='center', color='green', fontweight='bold')
for m in bot3_prom_mens:
    plt.scatter(m, prom_mens[m], color='red', s=100, zorder=5)
    plt.annotate(f"{prom_mens[m]:.1f}", (m, prom_mens[m]), textcoords="offset points",
                 xytext=(0,-15), ha='center', color='red', fontweight='bold')

plt.title(f"Régimen de Precipitación Mensual - {NOMBRE_SITIO}", fontsize=14, fontweight='bold')
plt.xlabel("Mes")
plt.ylabel("Precipitación Promedio (mm)")
plt.grid(True, linestyle='--', alpha=0.5)
plt.tight_layout()
plt.savefig(ruta_png_prom_mensual, dpi=150)
plt.close()
print(f"  Guardado: {ruta_png_prom_mensual}")

# ==========================================
# 7.1 HEATMAP AÑO vs MES
# ==========================================

print("Generando heatmap año vs mes...")

fig_heat, ax_heat = plt.subplots(figsize=(14, max(8, len(tabla) * 0.3)))

# Crear el heatmap con imshow
im = ax_heat.imshow(tabla.values, cmap='YlGnBu', aspect='auto')

# Configurar ejes
ax_heat.set_xticks(range(12))
ax_heat.set_xticklabels(meses_abbr, fontsize=10)
ax_heat.set_yticks(range(len(tabla)))
ax_heat.set_yticklabels(tabla.index, fontsize=9)

# Añadir valores en las celdas
for i in range(len(tabla)):
    for j in range(12):
        valor = tabla.iloc[i, j]
        if pd.notnull(valor):
            # Color del texto según el valor (blanco para valores altos, negro para bajos)
            color_texto = 'white' if valor > tabla.values[~np.isnan(tabla.values)].mean() * 1.3 else 'black'
            ax_heat.text(j, i, f'{valor:.0f}', ha='center', va='center',
                        fontsize=7, color=color_texto)

# Barra de colores
cbar = plt.colorbar(im, ax=ax_heat, shrink=0.8)
cbar.set_label('Precipitación (mm)', fontsize=10)

# Título y etiquetas
ax_heat.set_title(f"Mapa de Calor - Precipitación Mensual por Año\n{NOMBRE_SITIO}",
                  fontsize=14, fontweight='bold')
ax_heat.set_xlabel("Mes", fontsize=11)
ax_heat.set_ylabel("Año", fontsize=11)

# Líneas de separación
ax_heat.set_xticks(np.arange(-0.5, 12, 1), minor=True)
ax_heat.set_yticks(np.arange(-0.5, len(tabla), 1), minor=True)
ax_heat.grid(which='minor', color='white', linestyle='-', linewidth=0.5)

plt.tight_layout()
plt.savefig(ruta_png_heatmap, dpi=150, bbox_inches='tight')
plt.close()
print(f"  Guardado: {ruta_png_heatmap}")

# ==========================================
# 8. ANÁLISIS MULTITEMPORAL (FILTRO NO-NIÑA)
# ==========================================

p75_historico = tabla.quantile(0.75, axis=0)

scores_anios = []
for año in tabla.index:
    suma_lluvia_limpia = 0
    meses_valiosos = []
    fila = tabla.loc[año]
    for i, m_nom in enumerate(meses_abbr):
        m_num = i + 1
        valor = fila[m_nom]
        umbral = p75_historico[m_nom]
        if pd.notnull(valor):
            if valor > umbral and not es_periodo_nina(año, m_num):
                suma_lluvia_limpia += valor
                meses_valiosos.append((m_nom, valor))
    if meses_valiosos:
        scores_anios.append({'año': año, 'score': suma_lluvia_limpia, 'meses': meses_valiosos})

scores_anios.sort(key=lambda x: x['score'], reverse=True)
top6_multitemporal = scores_anios[:6]

# ==========================================
# 8.1 GRÁFICA MESES HÚMEDOS SELECCIONADOS
# ==========================================

print("Generando gráfica de meses húmedos seleccionados...")

# Preparar datos para la gráfica
datos_grafica = []
for item in top6_multitemporal:
    año = item['año']
    for mes_nom, valor in item['meses']:
        datos_grafica.append({
            'año': año,
            'mes': mes_nom,
            'valor': valor,
            'etiqueta': f"{mes_nom}\n{año}"
        })

if datos_grafica:
    # Ordenar por valor descendente
    datos_grafica.sort(key=lambda x: x['valor'], reverse=True)

    # Tomar los top 15 para no saturar la gráfica
    datos_grafica = datos_grafica[:15]

    fig_humedos, ax_humedos = plt.subplots(figsize=(14, 7))

    etiquetas = [d['etiqueta'] for d in datos_grafica]
    valores = [d['valor'] for d in datos_grafica]
    años = [d['año'] for d in datos_grafica]

    # Asignar colores según valor usando la misma paleta del heatmap (YlGnBu)
    norm = plt.Normalize(vmin=min(valores), vmax=max(valores))
    cmap_barras = plt.cm.YlGnBu
    colores_barras_h = [cmap_barras(norm(v)) for v in valores]

    barras_h = ax_humedos.bar(range(len(etiquetas)), valores, color=colores_barras_h, edgecolor='gray', linewidth=0.5)

    # Etiquetas de valores en las barras
    for i, (barra, valor) in enumerate(zip(barras_h, valores)):
        ax_humedos.annotate(f'{valor:.0f}',
                            xy=(barra.get_x() + barra.get_width()/2, valor),
                            xytext=(0, 3), textcoords='offset points',
                            ha='center', fontsize=8, fontweight='bold')

    # Línea de P75 histórico general
    p75_general = df['VALOR'].quantile(0.75)
    ax_humedos.axhline(y=p75_general, color='red', linestyle='--', linewidth=2,
                       label=f'P75 histórico: {p75_general:.1f} mm')

    ax_humedos.set_xticks(range(len(etiquetas)))
    ax_humedos.set_xticklabels(etiquetas, fontsize=9)

    # Leyenda con escala de color y P75
    leyenda_items = [
        mpatches.Patch(facecolor=cmap_barras(1.0), edgecolor='gray', label='Mayor precipitación'),
        mpatches.Patch(facecolor=cmap_barras(0.0), edgecolor='gray', label='Menor precipitación'),
        plt.Line2D([0], [0], color='red', linestyle='--', linewidth=2, label=f'P75: {p75_general:.1f} mm')
    ]
    ax_humedos.legend(handles=leyenda_items, loc='upper right', fontsize=9, framealpha=0.9)

    ax_humedos.set_title(f"Meses más Húmedos - {NOMBRE_SITIO}",
                         fontsize=13, fontweight='bold')
    ax_humedos.set_xlabel("Mes / Año")
    ax_humedos.set_ylabel("Precipitación (mm)")
    ax_humedos.grid(axis='y', linestyle='--', alpha=0.5)

    plt.tight_layout()
    plt.savefig(ruta_png_meses_humedos, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  Guardado: {ruta_png_meses_humedos}")
else:
    print("  No hay datos suficientes para la gráfica de meses húmedos")

# ==========================================
# 9. REPORTE TXT UNIFICADO
# ==========================================

print("Generando reporte de análisis completo...")

with open(ruta_analisis, "w", encoding="utf-8") as f:
    f.write("="*80 + "\n")
    f.write(f"ANÁLISIS COMPLETO DE PRECIPITACIÓN SATELITAL (TRMM/GPM IMERG) - {NOMBRE_SITIO}\n")
    f.write(f"Coordenadas: Lat {LATITUD}, Lon {LONGITUD}\n")
    f.write("="*80 + "\n\n")

    # 1. Información general
    f.write("1. INFORMACIÓN GENERAL DEL DATASET\n")
    f.write("-"*50 + "\n")
    f.write(f"Sitio: {NOMBRE_SITIO}\n")
    f.write(f"Fuente: Rasters TRMM y GPM IMERG\n")
    f.write(f"Período: {df['FECHA'].min().strftime('%Y-%m')} a {df['FECHA'].max().strftime('%Y-%m')}\n")
    f.write(f"Total registros: {len(df)} meses\n")
    f.write(f"Años completos: {len(df)//12}\n\n")

    # 2. Estadísticas básicas
    f.write("2. ESTADÍSTICAS BÁSICAS\n")
    f.write("-"*50 + "\n")
    f.write(f"Media mensual: {df['VALOR'].mean():.2f} mm\n")
    f.write(f"Mediana: {df['VALOR'].median():.2f} mm\n")
    f.write(f"Desviación estándar: {df['VALOR'].std():.2f} mm\n")
    f.write(f"Coeficiente de variación: {(df['VALOR'].std()/df['VALOR'].mean()*100):.1f}%\n")
    f.write(f"Mínimo: {df['VALOR'].min():.2f} mm ({df.loc[df['VALOR'].idxmin(), 'FECHA'].strftime('%Y-%m')})\n")
    f.write(f"Máximo: {df['VALOR'].max():.2f} mm ({df.loc[df['VALOR'].idxmax(), 'FECHA'].strftime('%Y-%m')})\n\n")

    # 3. Tendencia
    f.write("3. ANÁLISIS DE TENDENCIA\n")
    f.write("-"*50 + "\n")
    f.write(f"Tendencia: {'CRECIENTE' if pendiente > 0 else 'DECRECIENTE'}\n")
    f.write(f"Cambio por década: {cambio_decadal:.2f} mm/década\n\n")

    # 4. Régimen mensual → exportado a Excel (ver archivo _regimen_pluviometrico.xlsx)
    f.write("4. RÉGIMEN PLUVIOMÉTRICO MENSUAL\n")
    f.write("-"*50 + "\n")
    f.write(f"  (Ver tabla completa en: {os.path.basename(ruta_excel_regimen)})\n")
    f.write(f"\nMeses húmedos (> media): {', '.join([meses_nombres[m-1] for m in meses_humedos])}\n")
    f.write(f"Meses secos (<= media): {', '.join([meses_nombres[m-1] for m in meses_secos])}\n\n")

    # 5. Valores extremos
    f.write("5. 10 MESES MÁS LLUVIOSOS\n")
    f.write("-"*50 + "\n")
    top10 = df.nlargest(10, 'VALOR')[['FECHA', 'VALOR', 'ANOMALIA_PCT']]
    for i, (_, row) in enumerate(top10.iterrows(), 1):
        f.write(f"{i:>2}. {row['FECHA'].strftime('%Y-%m')}: {row['VALOR']:.1f} mm ({row['ANOMALIA_PCT']:+.1f}%)\n")

    f.write("\n6. 10 MESES MÁS SECOS\n")
    f.write("-"*50 + "\n")
    bot10 = df.nsmallest(10, 'VALOR')[['FECHA', 'VALOR', 'ANOMALIA_PCT']]
    for i, (_, row) in enumerate(bot10.iterrows(), 1):
        f.write(f"{i:>2}. {row['FECHA'].strftime('%Y-%m')}: {row['VALOR']:.1f} mm ({row['ANOMALIA_PCT']:+.1f}%)\n")

    # 7. Años extremos
    f.write("\n7. AÑOS EXTREMOS\n")
    f.write("-"*50 + "\n")
    f.write(f"Años más secos (< P10): {', '.join(map(str, años_secos))}\n")
    f.write(f"Años más húmedos (> P90): {', '.join(map(str, años_humedos))}\n\n")

    # 8. Análisis por décadas
    f.write("8. ANÁLISIS POR DÉCADAS\n")
    f.write("-"*50 + "\n")
    f.write(f"{'Década':<10} {'Media':>10} {'Total':>12} {'Registros':>10}\n")
    for decada, s in estadisticas_decadas.iterrows():
        f.write(f"{int(decada)}s{'':<6} {s['mean']:>10.1f} {s['sum']:>12.0f} {int(s['count']):>10}\n")

    # 9. Análisis multitemporal
    f.write("\n9. ANÁLISIS MULTITEMPORAL (FILTRO NO-NIÑA)\n")
    f.write("-"*50 + "\n")
    f.write("Años con mayor precipitación 'limpia' (> P75 y sin La Niña):\n\n")

    for i, item in enumerate(top6_multitemporal, 1):
        año = item['año']
        meses_info = sorted(item['meses'], key=lambda x: meses_abbr.index(x[0]))
        lista_meses = ", ".join([f"{m[0]} ({m[1]:.1f} mm)" for m in meses_info])
        f.write(f"  {i}. Año {año}: {lista_meses}\n")

    # 10. Recomendaciones
    f.write("\n10. RECOMENDACIONES PARA ANÁLISIS MULTITEMPORAL\n")
    f.write("-"*50 + "\n")
    f.write("Meses recomendados para análisis de inundaciones:\n")
    f.write(f"  {', '.join([meses_nombres[m-1] for m in meses_humedos[:4]])}\n\n")

    f.write("Años recomendados (sin influencia La Niña):\n")
    for item in top6_multitemporal[:3]:
        f.write(f"  - {item['año']}\n")

    f.write("\n" + "="*80 + "\n")
    f.write("FIN DEL ANÁLISIS\n")
    f.write("="*80 + "\n")

print(f"  Guardado: {ruta_analisis}")

# ==========================================
# 9b. RÉGIMEN PLUVIOMÉTRICO MENSUAL → EXCEL
# ==========================================

print("Generando Excel de régimen pluviométrico mensual...")

wb = Workbook()
ws = wb.active
ws.title = "Régimen Mensual"

# --- Estilos ---
fill_titulo   = PatternFill("solid", fgColor="1E6B1E")   # verde oscuro
fill_encab    = PatternFill("solid", fgColor="4CAF50")   # verde medio
fill_par      = PatternFill("solid", fgColor="C8E6C9")   # verde muy claro
fill_impar    = PatternFill("solid", fgColor="FFFFFF")   # blanco
fill_footer   = PatternFill("solid", fgColor="A5D6A7")   # verde claro

font_titulo   = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
font_encab    = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
font_datos    = Font(name="Calibri", size=11)
font_footer   = Font(name="Calibri", bold=True, size=10, color="1E6B1E")

borde_fino    = Border(
    left   = Side(style="thin", color="4CAF50"),
    right  = Side(style="thin", color="4CAF50"),
    top    = Side(style="thin", color="4CAF50"),
    bottom = Side(style="thin", color="4CAF50"),
)
alinear_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
alinear_izq    = Alignment(horizontal="left",   vertical="center")

num_cols = 6   # Mes, Media, Desv, CV%, Mín, Máx

# --- Fila 1: título ---
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
cel_tit = ws.cell(row=1, column=1,
                  value=f"RÉGIMEN PLUVIOMÉTRICO MENSUAL — {NOMBRE_SITIO.strip()}")
cel_tit.fill      = fill_titulo
cel_tit.font      = font_titulo
cel_tit.alignment = alinear_centro
ws.row_dimensions[1].height = 30

# --- Fila 2: subtítulo coordinadas ---
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
cel_sub = ws.cell(row=2, column=1,
                  value=f"Lat {LATITUD}  |  Lon {LONGITUD}  |  Fuente: TRMM/GPM IMERG")
cel_sub.fill      = PatternFill("solid", fgColor="2E7D32")
cel_sub.font      = Font(name="Calibri", size=10, color="FFFFFF")
cel_sub.alignment = alinear_centro
ws.row_dimensions[2].height = 18

# --- Fila 3: encabezados de columnas ---
encabezados = ["Mes", "Media (mm)", "Desv. Est.", "CV (%)", "Mín (mm)", "Máx (mm)"]
for col, texto in enumerate(encabezados, start=1):
    c = ws.cell(row=3, column=col, value=texto)
    c.fill      = fill_encab
    c.font      = font_encab
    c.alignment = alinear_centro
    c.border    = borde_fino
ws.row_dimensions[3].height = 22

# --- Filas de datos (filas 4-15) ---
for i, mes in enumerate(range(1, 13)):
    fila_excel = 4 + i
    fill_fila  = fill_par if i % 2 == 0 else fill_impar
    if mes in estadisticas_mensuales.index:
        s = estadisticas_mensuales.loc[mes]
        valores_fila = [
            meses_nombres[mes - 1],
            round(float(s['mean']), 1),
            round(float(s['std']),  1),
            round(float(s['cv']),   1),
            round(float(s['min']),  1),
            round(float(s['max']),  1),
        ]
    else:
        valores_fila = [meses_nombres[mes - 1], "-", "-", "-", "-", "-"]

    for col, valor in enumerate(valores_fila, start=1):
        c = ws.cell(row=fila_excel, column=col, value=valor)
        c.fill      = fill_fila
        c.font      = font_datos
        c.alignment = alinear_centro if col > 1 else alinear_izq
        c.border    = borde_fino
    ws.row_dimensions[fila_excel].height = 18

# --- Fila 16: espacio vacío ---
ws.row_dimensions[16].height = 8

# --- Fila 17: meses húmedos ---
ws.merge_cells(start_row=17, start_column=1, end_row=17, end_column=num_cols)
cel_hum = ws.cell(row=17, column=1,
                  value=f"Meses húmedos (> media): {', '.join([meses_nombres[m-1] for m in meses_humedos])}")
cel_hum.fill      = fill_footer
cel_hum.font      = font_footer
cel_hum.alignment = alinear_izq
ws.row_dimensions[17].height = 18

# --- Fila 18: meses secos ---
ws.merge_cells(start_row=18, start_column=1, end_row=18, end_column=num_cols)
cel_sec = ws.cell(row=18, column=1,
                  value=f"Meses secos (≤ media): {', '.join([meses_nombres[m-1] for m in meses_secos])}")
cel_sec.fill      = fill_footer
cel_sec.font      = font_footer
cel_sec.alignment = alinear_izq
ws.row_dimensions[18].height = 18

# --- Fila 20: sección ANÁLISIS DE TENDENCIA ---
ws.row_dimensions[19].height = 8  # espacio

ws.merge_cells(start_row=20, start_column=1, end_row=20, end_column=num_cols)
cel_tend_tit = ws.cell(row=20, column=1, value="ANÁLISIS DE TENDENCIA")
cel_tend_tit.fill      = fill_encab
cel_tend_tit.font      = font_encab
cel_tend_tit.alignment = alinear_izq
ws.row_dimensions[20].height = 20

tend_datos = [
    ("Tendencia:",          "CRECIENTE" if pendiente > 0 else "DECRECIENTE"),
    ("Cambio por década:",  f"{cambio_decadal:.2f} mm/década"),
]
for i, (etiqueta, valor) in enumerate(tend_datos):
    fila_e    = 21 + i
    fill_fila = fill_par if i % 2 == 0 else fill_impar
    c_et = ws.cell(row=fila_e, column=1, value=etiqueta)
    c_et.fill = fill_fila
    c_et.font = Font(name="Calibri", bold=True, size=11)
    c_et.alignment = alinear_izq
    c_et.border = borde_fino
    ws.merge_cells(start_row=fila_e, start_column=2, end_row=fila_e, end_column=num_cols)
    c_val = ws.cell(row=fila_e, column=2, value=valor)
    c_val.fill = fill_fila
    c_val.font = font_datos
    c_val.alignment = alinear_izq
    c_val.border = borde_fino
    ws.row_dimensions[fila_e].height = 18

# --- Fila 24: sección AÑOS EXTREMOS ---
ws.row_dimensions[23].height = 8  # espacio

ws.merge_cells(start_row=24, start_column=1, end_row=24, end_column=num_cols)
cel_ext_tit = ws.cell(row=24, column=1, value="AÑOS EXTREMOS")
cel_ext_tit.fill      = fill_encab
cel_ext_tit.font      = font_encab
cel_ext_tit.alignment = alinear_izq
ws.row_dimensions[24].height = 20

ext_datos = [
    ("Años más secos (< P10):",   ", ".join(map(str, años_secos))),
    ("Años más húmedos (> P90):", ", ".join(map(str, años_humedos))),
]
for i, (etiqueta, valor) in enumerate(ext_datos):
    fila_e    = 25 + i
    fill_fila = fill_par if i % 2 == 0 else fill_impar
    c_et = ws.cell(row=fila_e, column=1, value=etiqueta)
    c_et.fill = fill_fila
    c_et.font = Font(name="Calibri", bold=True, size=11)
    c_et.alignment = alinear_izq
    c_et.border = borde_fino
    ws.merge_cells(start_row=fila_e, start_column=2, end_row=fila_e, end_column=num_cols)
    c_val = ws.cell(row=fila_e, column=2, value=valor)
    c_val.fill = fill_fila
    c_val.font = font_datos
    c_val.alignment = alinear_izq
    c_val.border = borde_fino
    ws.row_dimensions[fila_e].height = 18

# --- Anchos de columna ---
anchos = [26, 14, 14, 12, 14, 14]
for col, ancho in enumerate(anchos, start=1):
    ws.column_dimensions[get_column_letter(col)].width = ancho

wb.save(ruta_excel_regimen)
print(f"  Guardado: {ruta_excel_regimen}")


# ==========================================
# 11. ANÁLISIS COMBINADO TRMM + ESTACIÓN IDEAM
# ==========================================

print("\n" + "="*60)
print(f"ANÁLISIS COMBINADO TRMM + ESTACIÓN IDEAM ({TIPO_ESTACION_IDEAM.upper()})")
print("="*60)

from collections import Counter

# RUTA_IDEAM_TXT definida en la sección 0 (CONFIGURACIÓN INICIAL)
ruta_analisis_combinado = os.path.join(RUTA_SALIDA, f"{nombre_base}_ANALISIS_COMBINADO.txt")

# Mapeos mes nombre <-> número (para parsear el TXT IDEAM)
_meses_map = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
              'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}
_meses_rev = {v: k for k, v in _meses_map.items()}

# ---- Parsear TXT del IDEAM ----
ideam = {'meses_humedos': [], 'prom_mensual': {}, 'anios_humedos': [], 'top10': []}
ideam_ok = False

# Variables z-score para limnimétrica (se definen aquí para evitar NameError en el reporte)
_trmm_z = pd.Series(dtype=float)
_lini_z = pd.Series(dtype=float)

try:
    with open(RUTA_IDEAM_TXT, 'r', encoding='utf-8') as _f:
        _lineas = _f.readlines()

    _seccion = None
    for _linea in _lineas:
        _ls = _linea.strip()

        # Detectar cambio de sección
        if '4. RÉGIMEN PLUVIOMÉTRICO MENSUAL' in _linea:
            _seccion = 'regimen'; continue
        elif '5. 10 VALORES MÁS ALTOS' in _linea:
            _seccion = 'top10'; continue
        elif '7. AÑOS EXTREMOS' in _linea:
            _seccion = 'extremos'; continue
        elif (_ls and _ls[0].isdigit() and len(_ls) > 1 and _ls[1] == '.'
              and _seccion == 'regimen'):
            _seccion = None

        # --- Sección régimen mensual ---
        if _seccion == 'regimen':
            _p = _ls.split('\t')
            if _p[0] in _meses_map and len(_p) >= 2:
                try:
                    ideam['prom_mensual'][_meses_map[_p[0]]] = float(_p[1])
                except ValueError:
                    pass
            if 'Meses húmedos' in _linea and ':' in _linea:
                _parte = _linea.split(':')[-1].strip()
                ideam['meses_humedos'] = [
                    _meses_map[m.strip()] for m in _parte.split(',')
                    if m.strip() in _meses_map
                ]

        # --- Sección top10 ---
        elif _seccion == 'top10':
            _p = _ls.split('\t')
            if len(_p) >= 3 and _p[0].isdigit():
                try:
                    _at = int(_p[1][:4]); _mt = int(_p[1][5:7])
                    ideam['top10'].append((_at, _mt, float(_p[2])))
                except (ValueError, IndexError):
                    pass

        # --- Sección años extremos ---
        elif _seccion == 'extremos' and 'más húmedos' in _linea and ':' in _linea:
            _parte = _linea.split(':')[-1].strip()
            ideam['anios_humedos'] = [
                int(a.strip()) for a in _parte.split(',') if a.strip().isdigit()
            ]

    ideam_ok = True
    print(f"  Meses húmedos IDEAM  : {[_meses_rev[m] for m in ideam['meses_humedos']]}")
    print(f"  Años P90 IDEAM       : {ideam['anios_humedos']}")
    print(f"  Top10 IDEAM parseados: {len(ideam['top10'])} registros")

except FileNotFoundError:
    print(f"  ADVERTENCIA: No se encontró el TXT IDEAM en:\n  {RUTA_IDEAM_TXT}")
    if TIPO_ESTACION_IDEAM == 'linimetrica':
        print("  Para estación limnimétrica se continuará solo con el CSV (TXT no requerido).")
        ideam_ok = True  # El CSV es suficiente para el análisis limnimétrico
    else:
        print("  El análisis combinado no se generará.")

if ideam_ok:

    # ---- Cargar CSV IDEAM (datos brutos: AÑO, MES, VALOR) ----
    _ideam_df = None
    try:
        _ideam_df = pd.read_csv(RUTA_IDEAM_CSV, encoding='latin-1', sep=None, engine='python')
        _ideam_df.columns = [c.strip().upper().replace('Ñ','N').replace('Á','A')
                              .replace('É','E').replace('Í','I').replace('Ó','O')
                              .replace('Ú','U') for c in _ideam_df.columns]
        _col_anio = next((c for c in _ideam_df.columns if 'A' in c and 'O' in c and len(c) <= 4), None)
        _col_mes  = next((c for c in _ideam_df.columns if c in ('MES','MES ')), None)
        _col_val  = next((c for c in _ideam_df.columns if 'VALOR' in c or 'PREC' in c), None)
        if not (_col_anio and _col_mes and _col_val):
            _col_anio, _col_mes, _col_val = _ideam_df.columns[-3], _ideam_df.columns[-2], _ideam_df.columns[-1]
        _ideam_df = _ideam_df[[_col_anio, _col_mes, _col_val]].copy()
        _ideam_df.columns = ['AÑO', 'MES', 'VALOR']
        _ideam_df['VALOR'] = pd.to_numeric(_ideam_df['VALOR'], errors='coerce')
        _ideam_df['AÑO']   = _ideam_df['AÑO'].astype(int)
        _ideam_df['MES']   = _ideam_df['MES'].astype(int)
        _ideam_df = _ideam_df.dropna(subset=['VALOR'])
        _ideam_df = _ideam_df[_ideam_df['AÑO'] >= 1985].copy()
        print(f"  CSV IDEAM cargado: {len(_ideam_df)} registros "
              f"({_ideam_df['AÑO'].min()}–{_ideam_df['AÑO'].max()}) [desde 1985]")

        # Para estación limnimétrica: derivar meses de nivel alto desde el CSV
        # si el TXT no aportó esa información (o no existía)
        if TIPO_ESTACION_IDEAM == 'linimetrica' and not ideam['meses_humedos']:
            _lini_prom_mens    = _ideam_df.groupby('MES')['VALOR'].mean()
            _lini_media_global = _lini_prom_mens.mean()
            ideam['meses_humedos'] = sorted(
                [m for m in _lini_prom_mens.index if _lini_prom_mens[m] > _lini_media_global]
            )
            ideam['prom_mensual'] = _lini_prom_mens.to_dict()
            print(f"  Meses nivel alto (linim.): "
                  f"{[_meses_rev.get(m, str(m)) for m in ideam['meses_humedos']]}")

    except Exception as _e:
        print(f"  No se pudo cargar CSV IDEAM: {_e}")

    # ---- 1. Consenso de meses húmedos / nivel alto ----
    _mc = sorted(set(meses_humedos) & set(ideam['meses_humedos']))
    _mt = sorted(set(meses_humedos) - set(ideam['meses_humedos']))
    _mi = sorted(set(ideam['meses_humedos']) - set(meses_humedos))

    # ---- 2. Score combinado ----
    _trmm_suma  = estadisticas_anuales['sum']
    _trmm_media = _trmm_suma.mean()
    _trmm_std   = _trmm_suma.std() if _trmm_suma.std() > 0 else 1

    _score      = {}
    _regla_anio = {}

    if _ideam_df is not None:
        _ideam_suma  = _ideam_df.groupby('AÑO')['VALOR'].sum()
        _ideam_media = _ideam_suma.mean()
        _ideam_std   = _ideam_suma.std() if _ideam_suma.std() > 0 else 1

        if TIPO_ESTACION_IDEAM == 'linimetrica':
            # ── Scoring normalizado (z-scores) para estación limnimétrica ──
            # Los valores no son comparables en unidades absolutas,
            # pero sí en comportamiento relativo respecto a la propia media histórica.
            _trmm_z = (_trmm_suma  - _trmm_media) / _trmm_std
            _lini_z = (_ideam_suma - _ideam_media) / _ideam_std

            # Incluir TODOS los años con datos: TRMM (1998+) y/o limnimétricos (cualquier año)
            _todos_anios = sorted(set(_trmm_suma.index) | set(_ideam_suma.index))
            for _a in _todos_anios:
                _has_trmm = _a in _trmm_z.index
                _has_lini = _a in _lini_z.index
                if _has_trmm and _has_lini:
                    _tz = float(_trmm_z[_a])
                    _lz = float(_lini_z[_a])
                    if _tz > 0 and _lz > 0:
                        _score[_a]      = (_tz + _lz) / 2
                        _regla_anio[_a] = 'R1: TRMM y linim. sobre media propia'
                    elif _tz > 0:
                        _score[_a]      = _tz * 0.75
                        _regla_anio[_a] = 'R2: solo TRMM sobre media'
                    elif _lz > 0:
                        _score[_a]      = _lz * 0.75
                        _regla_anio[_a] = 'R3: solo linim. sobre media'
                    else:
                        _score[_a]      = (_tz + _lz) / 2 * 0.3
                        _regla_anio[_a] = 'bajo en ambas fuentes'
                elif _has_lini:  # Pre-TRMM (antes de 1998) o sin cobertura satelital
                    _lz = float(_lini_z[_a])
                    _score[_a]      = _lz * 0.6
                    _regla_anio[_a] = 'solo linim. (pre-TRMM / sin cobertura satelital)'
                elif _has_trmm:  # Solo TRMM, fuera del periodo de la estación
                    _tz = float(_trmm_z[_a])
                    _score[_a]      = _tz * 0.5
                    _regla_anio[_a] = 'solo TRMM (sin dato estación)'

        else:  # ── pluviometrica: lógica original de comparación directa ──
            for _a in _trmm_suma.index:
                _t      = _trmm_suma[_a]
                _t_norm = (_t - _trmm_media) / _trmm_std

                if _a in _ideam_suma.index:
                    _i      = _ideam_suma[_a]
                    _i_norm = (_i - _ideam_media) / _ideam_std

                    _t_sobre_propia = _t > _trmm_media
                    _i_sobre_propia = _i > _ideam_media
                    _t_sobre_ideam  = _t > _ideam_media
                    _i_sobre_trmm   = _i > _trmm_media

                    if _t_sobre_propia and _i_sobre_propia:
                        _score[_a]      = (_t_norm + _i_norm) / 2
                        _regla_anio[_a] = 'R1: ambos > media propia'
                    elif _t_sobre_propia and _i_sobre_trmm:
                        _score[_a]      = (_t_norm + _i_norm) / 2 * 0.75
                        _regla_anio[_a] = 'R2: TRMM>media TRMM, IDEAM>media TRMM'
                    elif _i_sobre_propia and _t_sobre_ideam:
                        _score[_a]      = (_t_norm + _i_norm) / 2 * 0.75
                        _regla_anio[_a] = 'R3: IDEAM>media IDEAM, TRMM>media IDEAM'
                    else:
                        _score[_a]      = (_t_norm + _i_norm) / 2 * 0.3
                        _regla_anio[_a] = 'bajo en ambas fuentes'
                else:
                    _score[_a]      = _t_norm * 0.5
                    _regla_anio[_a] = 'solo TRMM (sin dato estación)'
    else:
        # Fallback: scoring solo con TRMM si no hay CSV
        for _a in _trmm_suma.index:
            _score[_a]      = (_trmm_suma[_a] - _trmm_media) / _trmm_std
            _regla_anio[_a] = 'solo TRMM (CSV no disponible)'

    _top10 = sorted(_score, key=_score.get, reverse=True)[:10]

    # Normativa IDEAM 2018: al menos 1 año anterior al 2000
    _anio_reemplazado = None
    _anio_incorporado = None
    if not any(a < 2000 for a in _top10):
        _pre2000 = [a for a in _score if a < 2000]
        if _pre2000:
            _mejor_pre2000    = max(_pre2000, key=lambda a: _score.get(a, 0))
            _anio_reemplazado = min(_top10, key=lambda a: _score.get(a, 0))
            _top10.remove(_anio_reemplazado)
            _top10.append(_mejor_pre2000)
            _anio_incorporado = _mejor_pre2000
            print(f"  Normativa IDEAM 2018: año {_anio_reemplazado} reemplazado por {_mejor_pre2000} (pre-2000)")
        else:
            print("  AVISO: No hay años anteriores al 2000 disponibles")

    # Rango reciente 2020-2026: al menos 2 años de ese período
    _anios_reemplazados_rec = []
    _anios_incorporados_rec = []
    _recientes_en_top = [a for a in _top10 if 2020 <= a <= 2026]
    _faltan_rec = 2 - len(_recientes_en_top)
    if _faltan_rec > 0:
        _candidatos_rec = sorted(
            [a for a in _score if 2020 <= a <= 2026 and a not in _top10],
            key=lambda a: _score.get(a, 0), reverse=True
        )
        for _mejor_rec in _candidatos_rec[:_faltan_rec]:
            _no_tocar = {_anio_incorporado} if _anio_incorporado else set()
            _no_tocar |= set(_recientes_en_top) | set(_anios_incorporados_rec)
            _candidatos_salida = [a for a in _top10 if a not in _no_tocar]
            if not _candidatos_salida:
                print("  AVISO: No se puede incorporar año reciente sin afectar normativa pre-2000")
                break
            _peor = min(_candidatos_salida, key=lambda a: _score.get(a, 0))
            _top10.remove(_peor)
            _top10.append(_mejor_rec)
            _anios_reemplazados_rec.append(_peor)
            _anios_incorporados_rec.append(_mejor_rec)
            _recientes_en_top.append(_mejor_rec)
            print(f"  Rango 2020-2026: año {_peor} reemplazado por {_mejor_rec} (reciente)")
        if not _candidatos_rec:
            print("  AVISO: No hay años entre 2020-2026 disponibles en el dataset")

    _top10_crono = sorted(_top10)

    # ---- 3. Temporalidad: meses > P75 en TRMM para top10 ----
    _p75h      = tabla.quantile(0.75, axis=0)
    _conteo_mc = Counter()
    _detalle   = {}
    for _a in _top10:
        if _a not in tabla.index:
            _detalle[_a] = []; continue
        _fila = tabla.loc[_a]
        _act  = []
        for _i2, _mn2 in enumerate(meses_abbr):
            _v = _fila[_mn2]
            if pd.notnull(_v) and _v > _p75h[_mn2]:
                _act.append((_i2 + 1, _mn2))
                if (_i2 + 1) in _mc:
                    _conteo_mc[_i2 + 1] += 1
        _detalle[_a] = _act

    _mr = [m for m in _mc if _conteo_mc.get(m, 0) >= 5]
    if not _mr:
        _mr = _mc

    # ---- 4. Escribir TXT combinado ----
    _SEP = "=" * 80
    _sep = "-" * 50
    _lbl_estacion = "LIMNIMÉTRICA" if TIPO_ESTACION_IDEAM == 'linimetrica' else "PLUVIOMÉTRICA"
    _lbl_valor    = "Nivel (mm regla)" if TIPO_ESTACION_IDEAM == 'linimetrica' else "Precip. (mm)"

    with open(ruta_analisis_combinado, 'w', encoding='utf-8') as f:
        f.write(_SEP + "\n")
        f.write(f"ANÁLISIS COMBINADO: TRMM/GPM IMERG + ESTACIÓN {_lbl_estacion} IDEAM\n")
        f.write(f"Sitio: {NOMBRE_SITIO} | Lat {LATITUD}, Lon {LONGITUD}\n")
        f.write(_SEP + "\n\n")

        f.write("FUENTES DE INFORMACIÓN\n")
        f.write(_sep + "\n")
        f.write(f"  Satelital (TRMM/GPM IMERG)      : "
                f"{df['FECHA'].min().strftime('%Y-%m')} a {df['FECHA'].max().strftime('%Y-%m')}\n")
        if _ideam_df is not None:
            f.write(f"  Estación {_lbl_estacion} (IDEAM): "
                    f"{_ideam_df['AÑO'].min()}-01 a {_ideam_df['AÑO'].max()}-12\n")
        else:
            f.write(f"  Estación {_lbl_estacion} (IDEAM): datos CSV no disponibles\n")
        if TIPO_ESTACION_IDEAM == 'linimetrica':
            f.write("\n  NOTA METODOLÓGICA:\n")
            f.write("  La estación limnimétrica registra NIVELES en mm de regla limnimétrica,\n")
            f.write("  NO precipitación en mm. Los sistemas de medición NO son directamente\n")
            f.write("  comparables en unidades absolutas. El análisis se basa en anomalías\n")
            f.write("  normalizadas (z-scores) para identificar años y meses relativamente\n")
            f.write("  húmedos/de nivel alto en cada fuente de forma independiente.\n")
            f.write("  Los registros anteriores a 1998 (pre-TRMM) se evalúan exclusivamente\n")
            f.write("  con el dato limnimétrico (score con descuento del 40%).\n")
        f.write("\n")

        # 1. Meses en consenso
        f.write("1. TEMPORALIDAD EN CONSENSO\n")
        f.write(_sep + "\n")
        _lbl_col_estacion = ("Meses nivel alto (linim.)" if TIPO_ESTACION_IDEAM == 'linimetrica'
                             else "Meses húmedos IDEAM    ")
        f.write(f"  Meses húmedos TRMM          : {', '.join([meses_nombres[m-1] for m in meses_humedos])}\n")
        f.write(f"  {_lbl_col_estacion}: "
                f"{', '.join([_meses_rev.get(m, str(m)) for m in ideam['meses_humedos']])}\n\n")
        if _mc:
            f.write(f"  CONSENSO (ambas fuentes)    : {', '.join([meses_nombres[m-1] for m in _mc])}\n")
        else:
            f.write("  Sin meses en pleno consenso\n")
        if _mt:
            f.write(f"  Solo en TRMM                : {', '.join([meses_nombres[m-1] for m in _mt])}\n")
        if _mi:
            f.write(f"  Solo en estación            : "
                    f"{', '.join([_meses_rev.get(m, str(m)) for m in _mi])}\n")
        f.write("\n")

        # 2. Régimen mensual comparado
        f.write("2. RÉGIMEN MENSUAL COMPARADO\n")
        f.write(_sep + "\n")
        if TIPO_ESTACION_IDEAM == 'linimetrica':
            f.write("  (TRMM = precipitación mm  |  Estación = nivel limnimétrico mm — unidades distintas)\n")
        f.write(f"  {'Mes':<6} {'TRMM (mm)':>12} {_lbl_valor:>18}  Estado\n")
        f.write(f"  {'-'*65}\n")
        for _m in range(1, 13):
            _nom    = meses_nombres[_m - 1]
            _tv     = float(prom_mens.iloc[_m - 1])
            _iv     = ideam['prom_mensual'].get(_m, float('nan'))
            _iv_str = f"{_iv:.1f}" if not np.isnan(_iv) else "-"
            if _m in _mc:       _estado = "CONSENSO - ambas fuentes"
            elif _m in _mt:     _estado = "Alto solo TRMM"
            elif _m in _mi:     _estado = "Alto solo estación"
            else:               _estado = "Bajo en ambas fuentes"
            f.write(f"  {_nom:<6} {_tv:>12.1f} {_iv_str:>18}  {_estado}\n")
        f.write("\n")

        # 3. Ranking top10 años
        f.write("3. TOP 10 AÑOS EN CONSENSO\n")
        f.write(_sep + "\n")
        if TIPO_ESTACION_IDEAM == 'linimetrica':
            f.write("  Metodología (anomalías normalizadas z-score):\n")
            f.write("    R1 (mejor): z_TRMM > 0  Y  z_linim > 0  (ambos sobre media propia)\n")
            f.write("    R2        : solo z_TRMM > 0  (score x0.75)\n")
            f.write("    R3        : solo z_linim > 0  (score x0.75)\n")
            f.write("    Pre-TRMM  : solo dato limnimétrico (score x0.6)\n")
        else:
            f.write("  Metodología de selección (reglas de consenso):\n")
            f.write("    R1 (mejor): TRMM > media TRMM  Y  IDEAM > media IDEAM\n")
            f.write("    R2        : TRMM > media TRMM  Y  IDEAM > media TRMM\n")
            f.write("    R3        : IDEAM > media IDEAM Y  TRMM > media IDEAM\n")
        f.write("  Normativa: Guia Tecnica IDEAM 2018 - al menos 1 año anterior al 2000\n")
        f.write("  Cobertura reciente: al menos 2 años entre 2020-2026\n\n")
        if _anio_incorporado:
            f.write(f"  AJUSTE NORMATIVO: año {_anio_reemplazado} reemplazado por {_anio_incorporado} (pre-2000)\n")
        for _pr, _pi in zip(_anios_reemplazados_rec, _anios_incorporados_rec):
            f.write(f"  AJUSTE 2020-2026 : año {_pr} reemplazado por {_pi} (reciente)\n")
        if _anio_incorporado or _anios_incorporados_rec:
            f.write("\n")

        if TIPO_ESTACION_IDEAM == 'linimetrica':
            _lbl_c2, _lbl_c3 = 'z_TRMM', 'z_linim'
        else:
            _lbl_c2, _lbl_c3 = 'TRMM mm', 'IDEAM mm'

        f.write(f"  {'Pos':>4} {'Año':>6} {'Score':>8} {_lbl_c2:>10} {_lbl_c3:>10}  Regla aplicada\n")
        f.write(f"  {'-'*65}\n")
        for _pos, _a in enumerate(sorted(_top10, key=lambda x: _score.get(x, 0), reverse=True), 1):
            _notas = []
            if _a == _anio_incorporado:
                _notas.append("pre-2000 (norma)")
            if _a in _anios_incorporados_rec:
                _notas.append("2020-2026 (reciente)")
            _nota  = ", ".join(_notas)
            _regla = _regla_anio.get(_a, '-')

            if TIPO_ESTACION_IDEAM == 'linimetrica':
                _tz_val = float(_trmm_z[_a]) if _a in _trmm_z.index else float('nan')
                _lz_val = float(_lini_z[_a]) if _a in _lini_z.index else float('nan')
                _c2 = f"{_tz_val:.2f}" if not np.isnan(_tz_val) else "-"
                _c3 = f"{_lz_val:.2f}" if not np.isnan(_lz_val) else "-"
            else:
                _suma_t = estadisticas_anuales.loc[_a, 'sum'] if _a in estadisticas_anuales.index else float('nan')
                _suma_i = (_ideam_suma.get(_a, float('nan'))
                           if _ideam_df is not None and hasattr(_ideam_suma, 'get') else float('nan'))
                _c2 = f"{_suma_t:.0f}" if not np.isnan(_suma_t) else "-"
                _c3 = f"{_suma_i:.0f}" if not np.isnan(_suma_i) else "-"

            f.write(f"  {_pos:>4} {_a:>6} {_score[_a]:>8.2f} {_c2:>10} {_c3:>10}  {_regla}  {_nota}\n")
        f.write("\n")

        # 4. Detalle temporalidad por año
        f.write("4. MESES > P75 HISTORICO EN LOS 10 AÑOS SELECCIONADOS\n")
        f.write(_sep + "\n")
        f.write("  (* = mes en consenso con estación)\n\n")
        for _a in _top10_crono:
            _act = _detalle.get(_a, [])
            if _act:
                _parts = [f"{meses_nombres[_mn-1]}{'*' if _mn in _mc else ''}" for _mn, _ in _act]
                f.write(f"  {_a}: {', '.join(_parts)}\n")
            else:
                f.write(f"  {_a}: sin datos TRMM o sin meses > P75\n")
        f.write("\n  Frecuencia de meses-consenso en los 10 años:\n")
        for _m in _mc:
            _c = _conteo_mc.get(_m, 0)
            f.write(f"    {meses_nombres[_m-1]:<5}: {_c}/10  {'|' * _c}\n")
        f.write("\n")

        # 5. Recomendación final
        f.write("5. RECOMENDACION FINAL PARA SELECCION MULTITEMPORAL\n")
        f.write(_sep + "\n")
        f.write("  Cumplimiento normativa: Guia Tecnica IDEAM 2018 - periodo >= 15 años,\n")
        f.write("  al menos 1 año anterior al 2000 incluido en la seleccion.\n")
        f.write("  Cobertura reciente: al menos 2 años entre 2020-2026.\n\n")
        f.write("  10 años recomendados (orden cronológico):\n")
        f.write(f"    {', '.join(map(str, _top10_crono))}\n")
        if _anio_incorporado:
            f.write(f"    (año {_anio_incorporado} incluido por normativa pre-2000)\n")
        if _anios_incorporados_rec:
            f.write(f"    (años {', '.join(map(str, _anios_incorporados_rec))} incluidos por cobertura 2020-2026)\n")
        f.write("\n")
        f.write("  Temporalidad en consenso (ambas fuentes):\n")
        f.write(f"    {', '.join([meses_nombres[m-1] for m in _mc])}\n\n")
        f.write("  Meses de mayor recurrencia en esos años:\n")
        f.write(f"    {', '.join([meses_nombres[m-1] for m in _mr])}\n")
        f.write("\n" + _SEP + "\n")
        f.write("FIN DEL ANALISIS COMBINADO\n")
        f.write(_SEP + "\n")

    print(f"  Guardado: {ruta_analisis_combinado}")
    print(f"  10 años consenso : {', '.join(map(str, _top10_crono))}")
    print(f"  Meses consenso   : {', '.join([meses_nombres[m-1] for m in _mc])}")
    print(f"  Meses recurrentes: {', '.join([meses_nombres[m-1] for m in _mr])}")

    # ---- 5. Gráfica comparativa ----
    print("  Generando gráfica comparativa TRMM vs IDEAM...")
    ruta_png_comparacion = os.path.join(RUTA_SALIDA, f"{nombre_base}_comparacion_TRMM_IDEAM.png")

    if _ideam_df is not None and len(_ideam_df) > 0:

        _trmm_anual  = df.groupby('AÑO')['VALOR'].sum()
        _ideam_anual = _ideam_df.groupby('AÑO')['VALOR'].sum()

        if TIPO_ESTACION_IDEAM == 'linimetrica':
            # ── Gráfica doble eje Y para estación limnimétrica ──
            # Eje izquierdo: precipitación TRMM (mm) — desde 1998
            # Eje derecho  : nivel limnimétrico acumulado (mm regla) — período completo incluyendo pre-1998
            _color_trmm = '#2E75B6'
            _color_lini = '#C55A11'

            _anio_ini_trmm = int(_trmm_anual.index.min())
            _anio_fin_trmm = int(_trmm_anual.index.max())
            _anio_ini_lini = int(_ideam_anual.index.min())
            _anio_fin_lini = int(_ideam_anual.index.max())
            _todos_anos = sorted(set(_trmm_anual.index) | set(_ideam_anual.index))

            # Filtro por año de inicio de comparación
            _anio_ini_graf = ANIO_INICIO_COMPARACION if ANIO_INICIO_COMPARACION is not None else _todos_anos[0]
            _todos_anos = [a for a in _todos_anos if a >= _anio_ini_graf]
            _trmm_graf   = _trmm_anual[_trmm_anual.index >= _anio_ini_graf]
            _lini_graf   = _ideam_anual[_ideam_anual.index >= _anio_ini_graf]

            fig_comp, ax1_comp = plt.subplots(figsize=(20, 10))
            fig_comp.subplots_adjust(bottom=0.22)
            ax2_comp = ax1_comp.twinx()

            # Serie TRMM (eje izquierdo)
            _trmm_anos_ord = sorted(_trmm_graf.index)
            ax1_comp.plot(_trmm_anos_ord,
                          [_trmm_graf[a] for a in _trmm_anos_ord],
                          color=_color_trmm, linewidth=2, marker='o', markersize=4,
                          label='TRMM / GPM IMERG — precipitación (mm)', zorder=3)

            # Serie limnimétrica (eje derecho)
            _lini_anos_ord = sorted(_lini_graf.index)
            ax2_comp.plot(_lini_anos_ord,
                          [_lini_graf[a] for a in _lini_anos_ord],
                          color=_color_lini, linewidth=2, marker='s', markersize=4,
                          label='Estación limnimétrica IDEAM — nivel (mm regla)', zorder=3, alpha=0.9)

            # Sombreado del período pre-TRMM si existe dentro del rango graficado
            _pre_trmm_anos = [a for a in _lini_anos_ord if a < _anio_ini_trmm]
            if _pre_trmm_anos:
                ax1_comp.axvspan(_pre_trmm_anos[0] - 0.5, _anio_ini_trmm - 0.5,
                                 color='lightyellow', alpha=0.45, zorder=0,
                                 label=f'Período pre-TRMM (solo limnimétrica, {_pre_trmm_anos[0]}–{_anio_ini_trmm-1})')
                ax1_comp.text((_pre_trmm_anos[0] + _anio_ini_trmm - 1) / 2, 0.96,
                              'Solo limnimétrica\n(pre-TRMM)',
                              ha='center', va='top', fontsize=8.5, color='#7A6000',
                              style='italic', transform=ax1_comp.get_xaxis_transform())

            # Medias históricas de cada fuente (calculadas sobre todo el registro, no solo el rango graficado)
            _med_t = _trmm_anual.mean()
            _med_l = _ideam_anual.mean()
            ax1_comp.axhline(_med_t, color=_color_trmm, linestyle='--', linewidth=1.2, alpha=0.55,
                             label=f'Media TRMM: {_med_t:.0f} mm')
            ax2_comp.axhline(_med_l, color=_color_lini, linestyle='--', linewidth=1.2, alpha=0.55,
                             label=f'Media nivel: {_med_l:.0f} mm')

            # Bandas doradas + anotaciones para los 10 años seleccionados (solo los visibles)
            _top_visibles = [a for a in _top10_crono if a in _todos_anos]
            for _a in _top_visibles:
                ax1_comp.axvspan(_a - 0.4, _a + 0.4, color='gold', alpha=0.35, zorder=1)
                if _a in _trmm_graf.index:
                    _vy = _trmm_graf[_a]
                    ax1_comp.annotate(str(_a), xy=(_a, _vy),
                                      xytext=(0, 8), textcoords='offset points',
                                      ha='center', fontsize=8, fontweight='bold',
                                      color='#7B5800', rotation=90)

            # Etiquetas de ejes con color diferenciado
            ax1_comp.set_ylabel('Precipitación TRMM/GPM anual acumulada (mm)',
                                color=_color_trmm, fontsize=11)
            ax2_comp.set_ylabel('Nivel limnimétrico anual acumulado (mm regla)',
                                color=_color_lini, fontsize=11)
            ax1_comp.tick_params(axis='y', labelcolor=_color_trmm)
            ax2_comp.tick_params(axis='y', labelcolor=_color_lini)

            ax1_comp.set_xticks(_todos_anos[::2])
            ax1_comp.set_xticklabels(_todos_anos[::2], rotation=45, fontsize=8)
            ax1_comp.set_xlim(_todos_anos[0] - 0.8, _todos_anos[-1] + 0.8)
            ax1_comp.set_xlabel('Año', fontsize=11)

            # ── Convención: parche dorado para años seleccionados ──
            _parche_gold = mpatches.Patch(facecolor='gold', edgecolor='#7B5800',
                                          alpha=0.6, label=f'Años seleccionados ({len(_top10_crono)}): '
                                          + ', '.join(map(str, _top10_crono)))
            _lines1, _labs1 = ax1_comp.get_legend_handles_labels()
            _lines2, _labs2 = ax2_comp.get_legend_handles_labels()
            ax1_comp.legend(_lines1 + _lines2 + [_parche_gold],
                            _labs1 + _labs2 + [_parche_gold.get_label()],
                            fontsize=8.5, loc='upper center',
                            bbox_to_anchor=(0.5, -0.12),
                            framealpha=0.92, ncol=3, borderpad=0.7)

            ax1_comp.set_title(
                f'Comparación: TRMM/GPM IMERG (precipitación) vs Estación Limnimétrica (nivel)\n'
                f'{NOMBRE_SITIO}  |  TRMM: {_anio_ini_trmm}–{_anio_fin_trmm}  '
                f'|  Linim.: {_anio_ini_lini}–{_anio_fin_lini}  '
                f'|  Gráfica desde: {_anio_ini_graf}',
                fontsize=11, fontweight='bold')
            ax1_comp.grid(axis='y', linestyle='--', alpha=0.3)

            # ── Nota metodológica debajo del gráfico (fuera del área del eje) ──
            _nota_lini = (
                "Se seleccionan los años en que TRMM y la estación IDEAM coinciden en superar "
                "su propio promedio histórico, priorizando aquellos que además superan el percentil 75 de cada fuente."
            )
            fig_comp.text(0.5, 0.005, _nota_lini,
                          ha='center', va='bottom', fontsize=10.5, wrap=True,
                          bbox=dict(boxstyle='round,pad=0.5', facecolor='#FFFDE7',
                                    edgecolor='#C8A000', alpha=0.90),
                          transform=fig_comp.transFigure)

        else:  # ── pluviometrica: gráfica eje único ──
            _anio_ini_comun = max(_trmm_anual.index.min(), _ideam_anual.index.min())
            _anio_fin_comun = min(_trmm_anual.index.max(), _ideam_anual.index.max())

            # Filtro por año de inicio de comparación
            _anio_ini_graf = ANIO_INICIO_COMPARACION if ANIO_INICIO_COMPARACION is not None else _anio_ini_comun
            _anio_ini_graf = max(_anio_ini_graf, _anio_ini_comun)  # no puede ser anterior al período común

            _trmm_plot  = _trmm_anual[_trmm_anual.index.isin(range(_anio_ini_graf, _anio_fin_comun + 1))]
            _ideam_plot = _ideam_anual[_ideam_anual.index.isin(range(_anio_ini_graf, _anio_fin_comun + 1))]
            _anios_plot = sorted(set(_trmm_plot.index) | set(_ideam_plot.index))

            _color_trmm  = '#2E75B6'
            _color_ideam = '#C55A11'

            fig_comp, ax_comp = plt.subplots(figsize=(18, 10))
            fig_comp.subplots_adjust(bottom=0.22)

            ax_comp.plot(sorted(_trmm_plot.index),
                         [_trmm_plot[a] for a in sorted(_trmm_plot.index)],
                         color=_color_trmm, linewidth=2, marker='o', markersize=4,
                         label='TRMM / GPM IMERG (satelital)', zorder=3)
            ax_comp.plot(sorted(_ideam_plot.index),
                         [_ideam_plot[a] for a in sorted(_ideam_plot.index)],
                         color=_color_ideam, linewidth=2, marker='s', markersize=4,
                         label='Estación IDEAM (campo)', zorder=3)

            # Medias sobre todo el período común (no solo el rango graficado)
            _med_t = _trmm_anual[_trmm_anual.index.isin(range(_anio_ini_comun, _anio_fin_comun + 1))].mean()
            _med_i = _ideam_anual[_ideam_anual.index.isin(range(_anio_ini_comun, _anio_fin_comun + 1))].mean()
            ax_comp.axhline(_med_t, color=_color_trmm, linestyle='--', linewidth=1.2, alpha=0.6,
                            label=f'Media TRMM: {_med_t:.0f} mm')
            ax_comp.axhline(_med_i, color=_color_ideam, linestyle='--', linewidth=1.2, alpha=0.6,
                            label=f'Media IDEAM: {_med_i:.0f} mm')

            # Bandas doradas + anotaciones para los años visibles en el rango graficado
            _top_visibles = [a for a in _top10_crono if a in _anios_plot]
            for _a in _top_visibles:
                ax_comp.axvspan(_a - 0.4, _a + 0.4, color='gold', alpha=0.35, zorder=1)
                _vy = max(_trmm_plot.get(_a, 0), _ideam_plot.get(_a, 0))
                ax_comp.annotate(str(_a), xy=(_a, _vy),
                                 xytext=(0, 8), textcoords='offset points',
                                 ha='center', fontsize=8, fontweight='bold',
                                 color='#7B5800', rotation=90)

            ax_comp.set_xticks(_anios_plot[::2])
            ax_comp.set_xticklabels(_anios_plot[::2], rotation=45, fontsize=8)
            ax_comp.set_xlim(_anios_plot[0] - 0.8, _anios_plot[-1] + 0.8)
            ax_comp.set_xlabel('Año', fontsize=11)
            ax_comp.set_ylabel('Precipitación anual total (mm)', fontsize=11)

            # ── Convención: parche dorado para años seleccionados ──
            _parche_gold = mpatches.Patch(facecolor='gold', edgecolor='#7B5800',
                                          alpha=0.6, label=f'Años seleccionados ({len(_top10_crono)}): '
                                          + ', '.join(map(str, _top10_crono)))
            _handles, _labels = ax_comp.get_legend_handles_labels()
            ax_comp.legend(_handles + [_parche_gold], _labels + [_parche_gold.get_label()],
                           fontsize=8.5, loc='upper center',
                           bbox_to_anchor=(0.5, -0.12),
                           framealpha=0.92, ncol=3)

            ax_comp.set_title(
                f'Comparación Precipitación Anual: TRMM/GPM IMERG vs Estación IDEAM\n'
                f'{NOMBRE_SITIO}  |  Periodo común: {_anio_ini_comun}–{_anio_fin_comun}'
                f'  |  Gráfica desde: {_anio_ini_graf}',
                fontsize=12, fontweight='bold')
            ax_comp.grid(axis='y', linestyle='--', alpha=0.4)

            # ── Nota metodológica debajo del gráfico (fuera del área del eje) ──
            _nota_pluv = (
                "Se seleccionan los años en que TRMM y la estación IDEAM coinciden en superar "
                "su propio promedio histórico, priorizando aquellos que además superan el percentil 75 de cada fuente."
            )
            fig_comp.text(0.5, 0.005, _nota_pluv,
                          ha='center', va='bottom', fontsize=10.5, wrap=True,
                          bbox=dict(boxstyle='round,pad=0.5', facecolor='#FFFDE7',
                                    edgecolor='#C8A000', alpha=0.90),
                          transform=fig_comp.transFigure)

        plt.savefig(ruta_png_comparacion, dpi=150, bbox_inches='tight')
        plt.close(fig_comp)
        print(f"    Guardado: {ruta_png_comparacion}")
    else:
        print("    CSV IDEAM no disponible, gráfica comparativa omitida.")


# ==========================================
# 10. RESUMEN FINAL
# ==========================================

print("\n" + "="*60)
print(f"RESUMEN DEL ANÁLISIS - {NOMBRE_SITIO}")
print("="*60)
print(f"Precipitación media: {df['VALOR'].mean():.2f} mm")
print(f"Tendencia: {cambio_decadal:+.2f} mm/década")
print(f"Meses más lluviosos: {', '.join([meses_nombres[m-1] for m in meses_humedos[:3]])}")
print(f"Años recomendados: {', '.join([str(item['año']) for item in top6_multitemporal[:3]])}")
print("="*60)
print("\nARCHIVOS GENERADOS:")
print(f"  - CSV: {ruta_csv}")
print(f"  - Serie temporal: {ruta_png_serie}")
print(f"  - Tabla intranual: {ruta_png_intranual}")
print(f"  - Tabla interanual: {ruta_png_interanual}")
print(f"  - Promedio anual: {ruta_png_prom_anual}")
print(f"  - Promedio mensual: {ruta_png_prom_mensual}")
print(f"  - Heatmap año/mes: {ruta_png_heatmap}")
print(f"  - Meses húmedos seleccionados: {ruta_png_meses_humedos}")
print(f"  - Análisis completo: {ruta_analisis}")
print(f"  - Régimen mensual (Excel): {ruta_excel_regimen}")
if ideam_ok:
    print(f"  - Comparación TRMM+IDEAM: {ruta_png_comparacion}")
print("\nPROCESO FINALIZADO.")
